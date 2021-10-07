#!/usr/bin/env python3

import sys
from openpyxl import load_workbook
import json

try:
    xlsx_in = sys.argv[1]
    generic_part_file = sys.argv[2]
    xlsx_out = sys.argv[3]
except:
    print("Usage: " + argv[0] + " spreadsheet_in part_association.json spreadsheet_out")
    exit(1)

wb = load_workbook(xlsx_in)
ws = wb.active

generic_part_dict = json.load(open(generic_part_file))

for col in ws.iter_cols(min_row=7, min_col=10, max_col=10):
    for cell in col:
        lib_pn = ws['C'][cell.row - 1].value
        part = generic_part_dict.get(lib_pn)
        if part != None:
            pn = part["partnumber"]
            manu = part["manufacturer"]
            notes = part["notes"]
            ws['M'][cell.row - 1].value = pn
            ws['L'][cell.row - 1].value = manu
            ws['S'][cell.row - 1].value = notes

wb.save(xlsx_out)
